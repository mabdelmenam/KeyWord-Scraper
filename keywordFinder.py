import requests
import string
import json
from openpyxl import *
from openpyxl.styles import Font
from pathlib import Path

# Getting user input for amount of searches
numInputs = int(input("How many searches are you making?: "))
searchList = [None] * numInputs

for x in range(numInputs):
    searchQuery = input("Enter search " + str(x) + ": ")
    searchList[x] = searchQuery

print(searchList)

# Initializing variables for excel sheet
my_file = Path("Keywords.xlsx")
wb = None
sheet = None

# Checking if file exists
if(my_file.is_file()):
    wb = load_workbook("Keywords.xlsx")
    sheet = wb.active

else:  # If file does not exist, create it
    wb = Workbook()
    sheet = wb.active

    c = sheet.cell(row=1, column=1)
    c.value = "Results"
    ft = Font(size=12, bold=True)
    c.font = ft

row = sheet.max_row + 1  # Writing to the amount of rows + 1
column = 1

# Looping through search results
for i in range(len(searchList)):
    URL = "http://suggestqueries.google.com/complete/search?client=firefox&q=" + \
        searchList[i]
    headers = {'User-agent': 'Mozilla/5.0'}
    response = requests.get(URL, headers=headers)
    result = json.loads(response.text)
    for item in (result[1]):  # Writing to excel sheet
        c1 = sheet.cell(row=row, column=column)
        c1.value = item
        row += 1

wb.save('Keywords.xlsx')
#abc = string.ascii_lowercase
#abcSplit = list(abc)
# print(abcSplit)
print(result[1])
